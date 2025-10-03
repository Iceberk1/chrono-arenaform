from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
import sqlite3
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment

import io
import pandas as pd
from datetime import datetime

app = Flask(__name__)
DB_FILE = 'chrono_event.db'

# --- Database setup ---
def init_db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA journal_mode=WAL;")
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS candidates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            number INTEGER UNIQUE,
            first_name TEXT,
            last_name TEXT,
            email TEXT,
            phone TEXT,
            UNIQUE(first_name, last_name)
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_number INTEGER,
            circuit INTEGER,
            time REAL,
            touches INTEGER DEFAULT 0,
            created_at DATETIME DEFAULT (DATE('now'))
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# --- Routes ---
@app.route('/')
def index():
    return redirect(url_for('chrono'))

@app.route('/add_candidate', methods=['GET','POST'])
def add_candidate():
    if request.method == 'POST':
        data = request.get_json()
        no_contact = data.get('no_contact', False)

        # Validation seulement si on ne bypass pas
        if not no_contact:
            email_pattern = r'^\S+@\S+\.\S+$'
            phone_pattern = r'^\d{10}$'
            if not re.match(email_pattern, data.get('email', '')):
                return jsonify(success=False, error='Email mal formaté')
            if not re.match(phone_pattern, data.get('phone', '')):
                return jsonify(success=False, error='Téléphone mal formaté')

        conn = sqlite3.connect(DB_FILE)
        conn.execute("PRAGMA journal_mode=WAL;")
        c = conn.cursor()
        c.execute('SELECT COUNT(*) FROM candidates WHERE first_name=? AND last_name=?',
                  (data['first_name'], data['last_name']))
        if c.fetchone()[0] > 0:
            conn.close()
            return jsonify(success=False, error='Candidat déjà existant')

        c.execute('SELECT COALESCE(MAX(number),0)+1 FROM candidates')
        next_number = c.fetchone()[0]

        # Si bypass, mettre email et phone à None
        email = None if no_contact else data.get('email')
        phone = None if no_contact else data.get('phone')

        c.execute('INSERT INTO candidates (number, first_name, last_name, email, phone) VALUES (?,?,?,?,?)',
                  (next_number, data['first_name'], data['last_name'], email, phone))
        conn.commit()
        conn.close()
        return jsonify(success=True, number=next_number)
        pass
        
    # GET : récupération du message si un candidat vient d'être ajouté
    added_number = request.args.get('added')
    message = ''
    if added_number:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('SELECT first_name, last_name FROM candidates WHERE number=?', (added_number,))
        row = c.fetchone()
        conn.close()
        if row:
            first_name, last_name = row
            message = f"Candidat {first_name} {last_name} ajouté avec succès ! Numéro : {added_number}"

    return render_template("add_candidates.html", message=message)


@app.route('/chrono')
def chrono():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA journal_mode=WAL;")
    c = conn.cursor()
    c.execute('SELECT number, first_name, last_name FROM candidates ORDER BY number')
    candidates = c.fetchall()

    # message si temps enregistré
    saved_time = request.args.get('saved_time')
    circuit = request.args.get('circuit')
    message = ''
    if saved_time and circuit:
        c.execute("SELECT first_name, last_name FROM candidates WHERE number=?", (saved_time,))
        row = c.fetchone()
        if row:
            first_name, last_name = row
            message = f"Temps enregistré pour {first_name} {last_name} sur le circuit {circuit} ✅"

    conn.close()
    return render_template("chrono.html", candidates=candidates, message=message)


@app.route('/save_time', methods=['POST'])
def save_time():
    data = request.json
    number = data['number']
    circuit = int(data['circuit'])
    time = float(data['time'])
    touches = int(data.get('touches', 0))

    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA journal_mode=WAL;")
    c = conn.cursor()

    # Insérer la nouvelle valeur (nouvel id)
    c.execute('''
        INSERT INTO results (candidate_number, circuit, time, touches, created_at)
        VALUES (?, ?, ?, ?, DATE('now'))
    ''', (number, circuit, time, touches))

    conn.commit()
    conn.close()
    return redirect(url_for('chrono', saved_time=number, circuit=circuit))


@app.route('/results')
def results():
    date_str = request.args.get('date')

    if date_str:
        try:
            # Accepte format YYYY-MM-DD
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            # fallback : jour courant
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()

    display_date = selected_date.strftime("%d/%m/%Y")

    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA journal_mode=WAL;")
    c = conn.cursor()
    results_dict = {}

    # Dates dispo dans la base
    c.execute("SELECT DISTINCT DATE(created_at) FROM results ORDER BY DATE(created_at) DESC")
    available_dates = [row[0] for row in c.fetchall()]

    for circuit in range(1, 5):
        if circuit == 3:
            c.execute('''
                SELECT r.candidate_number, c.last_name, c.first_name,
                       printf("%02d:%02d.%02d", r.time/60, r.time%60, (r.time*100)%100),
                       r.touches
                FROM results r
                JOIN candidates c ON r.candidate_number = c.number
                WHERE r.circuit = ?
                  AND DATE(r.created_at) = ?
                ORDER BY r.touches DESC, r.time ASC
            ''', (circuit, selected_date))
        elif circuit == 4:
            c.execute('''
                SELECT r.candidate_number, c.last_name, c.first_name,
                       printf("%02d:%02d.%02d", r.time/60, r.time%60, (r.time*100)%100)
                FROM results r
                JOIN candidates c ON r.candidate_number = c.number
                WHERE r.circuit = ?
                  AND DATE(r.created_at) = ?
                ORDER BY r.time DESC
            ''', (circuit, selected_date))
        else:
            c.execute('''
                SELECT r.candidate_number, c.last_name, c.first_name,
                       printf("%02d:%02d.%02d", r.time/60, r.time%60, (r.time*100)%100)
                FROM results r
                JOIN candidates c ON r.candidate_number = c.number
                WHERE r.circuit = ?
                  AND DATE(r.created_at) = ?
                ORDER BY r.time ASC
            ''', (circuit, selected_date))

        rows = c.fetchall()

        # Dictionnaire pour stocker la meilleure perf par candidat
        unique_best = {}
        for r in rows:
            cand_number = r[0]
            if cand_number not in unique_best:
                unique_best[cand_number] = r  # première occurrence (meilleure car déjà trié)
            # sinon on ignore car une meilleure perf est déjà enregistrée

        # On garde uniquement les premiers candidats uniques
        best = list(unique_best.values())[:3]

        # Derniers 5
        c.execute(f'''
            SELECT r.candidate_number, c.last_name, c.first_name,
                   printf("%02d:%02d.%02d", r.time/60, r.time%60, (r.time*100)%100)
                   {', r.touches' if circuit==3 else ''}
            FROM results r
            JOIN candidates c ON r.candidate_number = c.number
            WHERE r.circuit = ?
              AND DATE(r.created_at) = ?
            ORDER BY r.id DESC
            LIMIT 5
        ''', (circuit, selected_date))
        last = c.fetchall()

        results_dict[circuit] = {'best': best, 'last': last}

    conn.close()
    return render_template(
        "results.html",
        results=results_dict,
        selected_date=selected_date,
        available_dates=available_dates,
        display_date=display_date
    )

# EXPORT excel
@app.route('/export_excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill, Font
    import io

    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA journal_mode=WAL;")
    c = conn.cursor()

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidats"

    # Styles Arenaform
    header_fill = PatternFill(start_color="D63138", end_color="D63138", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

    # En-têtes
    headers = ["Numéro", "Nom", "Prénom", "Email", "Téléphone",
               "Circuit 1 - Ninja", "Circuit 2 - Crossfit/Hyrox", "Circuit 3 - Précision (temps)", "Circuit 3 - Précision (touches)", "Circuit 4 - Suspension"]
    ws.append(headers)

    # Style en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Récupérer candidats
    c.execute("SELECT number, last_name, first_name, email, phone FROM candidates")
    candidates = c.fetchall()

    row_index = 2
    for cand in candidates:
        number = cand[0]
        row = list(cand)

        # Récup résultats circuits
        for circuit in range(1, 5):
            if circuit == 3:
                c.execute("SELECT time, touches FROM results WHERE candidate_number=? AND circuit=3", (number,))
                res = c.fetchall()
                if res:
                    times = "\n".join([f"{int(t[0]//60):02d}:{int(t[0]%60):02d}.{int(t[0]*100%100):02d}" for t in res])
                    touches = "\n".join([str(t[1]) for t in res])
                else:
                    times, touches = "", ""
                row.extend([times, touches])
            else:
                c.execute("SELECT time FROM results WHERE candidate_number=? AND circuit=?", (number, circuit))
                res = c.fetchall()
                if res:
                    times = "\n".join([f"{int(t[0]//60):02d}:{int(t[0]%60):02d}.{int(t[0]*100%100):02d}" for t in res])
                else:
                    times = ""
                row.append(times)

        ws.append(row)

        # Appliquer style à la ligne (alternance gris clair)
        if row_index % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row_index, col).fill = alt_fill

        row_index += 1

    conn.close()

    # Ajustement auto largeur colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(wrapText=True, vertical="top")
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)  # limite largeur max

    # Export
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="candidats.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Page de stats
@app.route("/stats")
def stats():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Récupérer toutes les dates disponibles (basées sur created_at)
    c.execute("SELECT DISTINCT DATE(created_at) FROM results ORDER BY created_at DESC")
    dates = [row[0] for row in c.fetchall()]

    # Récupérer la date sélectionnée (par défaut aujourd'hui)
    selected_date = request.args.get("date")
    if not selected_date:
        c.execute("SELECT DATE('now')")
        selected_date = c.fetchone()[0]

    # Stats par circuit pour la date sélectionnée
    stats_today = {}
    for circuit in range(1, 5):
        c.execute(
            "SELECT COUNT(*) FROM results WHERE circuit=? AND DATE(created_at)=?",
            (circuit, selected_date)
        )
        stats_today[circuit] = c.fetchone()[0]

    # Total du jour (tous circuits confondus)
    c.execute("SELECT COUNT(*) FROM results WHERE DATE(created_at)=?", (selected_date,))
    total_today = c.fetchone()[0]

    # Totaux all time
    stats_all = {}
    for circuit in range(1, 5):
        c.execute("SELECT COUNT(*) FROM results WHERE circuit=?", (circuit,))
        stats_all[circuit] = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM results")
    total_all = c.fetchone()[0]

    conn.close()

    return render_template(
        "stats.html",
        dates=dates,
        selected_date=selected_date,
        stats_today=stats_today,
        total_today=total_today,
        stats_all=stats_all,
        total_all=total_all
    )

    
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)